
# hunt.py - MASTER-LEVEL COVERS (no thumbnails)
# BUILD_ID: HUNT_20260109_COVERS_V4
#
# - Reads per-decade *_top_sellers_*.csv from data_backups
# - For each release_id, calls Discogs /releases/{id}
# # - Chooses a cover image (primary preferred)
# - Saves covers into D:\records\outputs\hunt\covers
# - Builds simple HTML pages to a single index.html (full-size covers + year/artist/title)
#
# NOTE: This script intentionally has no collection writes.

import csv
import io
import json
import logging
import os
import re
import sys
import time
from pathlib import Path
from typing import Dict, List, Tuple

import requests

try:
    import openpyxl  # only used when reading .xlsx fallback
except ImportError:
    openpyxl = None

# Load .env so DISCOGS_* vars are available
try:
    from dotenv import load_dotenv
    load_dotenv(dotenv_path=Path(r"D:\records\.env"))
except Exception:
    pass


# ---------- paths ----------

REPO_ROOT = Path(r"C:\Users\David\Documents\GitHub\code\records")
INPUT_DIR = REPO_ROOT / "data_backups"

RUNTIME_ROOT = Path(r"D:\records")
OUTPUT_DIR = RUNTIME_ROOT / "outputs" / "hunt"
COVERS_DIR = OUTPUT_DIR / "covers"


# ---------- logging ----------

def setup_logging() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log_path = OUTPUT_DIR / "hunt.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    logging.info("=== hunt.py starting (BUILD_ID=HUNT_20260109_COVERS_V4) ===")
    logging.info("  INPUT_DIR=%s", INPUT_DIR)
    logging.info("  OUTPUT_DIR=%s", OUTPUT_DIR)


# ---------- Discogs client ----------

TOKEN = os.getenv("DISCOGS_TOKEN", "").strip()
USER = os.getenv("DISCOGS_USERNAME", "").strip()
UA = os.getenv("DISCOGS_USER_AGENT", "").strip() or "HuntCovers/1.0 +https://whitepup.github.io/store/"


class Client:
    def __init__(self) -> None:
        self.s = requests.Session()
        self.s.headers.update(
            {
                "User-Agent": UA,
                "Authorization": f"Discogs token={TOKEN}",
            }
        )
        self.base = "https://api.discogs.com"

    def get_release(self, rid: int) -> Dict:
        url = f"{self.base}/releases/{rid}"
        for attempt in range(3):
            try:
                r = self.s.get(url, timeout=30)
            except Exception as e:
                logging.warning("get_release %s network error (%s)", rid, e)
                time.sleep(2)
                continue

            if r.status_code == 429:
                wait = int(r.headers.get("Retry-After", "5")) + 1
                logging.info("get_release %s 429 -> sleeping %ss", rid, wait)
                time.sleep(wait)
                continue

            if r.status_code >= 400:
                logging.warning("get_release %s HTTP %s %s", rid, r.status_code, r.text[:200])
                return {}

            try:
                return r.json()
            except Exception as e:
                logging.warning("get_release %s JSON error: %s", rid, e)
                return {}
        return {}


# ---------- helpers for input ----------

def infer_decade(p: Path) -> str:
    m = re.search(r"(19[0-9]{2}s|20[0-9]{2}s)", p.name)
    return m.group(1) if m else "unknown"


def find_input_files() -> List[Path]:
    if not INPUT_DIR.exists():
        logging.warning("INPUT_DIR does not exist: %s", INPUT_DIR)
        return []
    files: List[Path] = []
    files += sorted(INPUT_DIR.glob("*_top_sellers_*.xlsx"))
    files += sorted(INPUT_DIR.glob("*_top_sellers_*.csv"))
    logging.info("Found %d input files", len(files))
    return files


def read_xlsx(p: Path) -> List[Dict]:
    if openpyxl is None:
        logging.error("openpyxl not installed, cannot read %s", p)
        return []
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    head = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    out: List[Dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {str(head[i]): r[i] for i in range(len(head))}
        out.append(row)
    logging.info("  Loaded %d from %s", len(out), p.name)
    return out


def read_csv(p: Path) -> List[Dict]:
    out: List[Dict] = []
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            out.append(r)
    logging.info("  Loaded %d from %s", len(out), p.name)
    return out


def extract_release_id(row: Dict) -> int | None:
    for k in ("discogs_id", "release_id", "id", "release"):
        if k in row and row[k] not in (None, ""):
            try:
                return int(str(row[k]).strip())
            except Exception:
                pass
    return None



def choose_cover(rel: Dict) -> str:
    """
    Choose a cover image URL for a release.

    Streamlined behavior (no thumbnails, no extra network requests):
      - Prefer the image with type == "primary" when present
      - Otherwise fall back to the first image
      - Return the full-size URL (uri or resource_url)
    """
    images = rel.get("images") or []
    if not images:
        return ""

    for img in images:
        if str(img.get("type", "")).lower() == "primary":
            return img.get("uri") or img.get("resource_url") or ""

    img = images[0]
    return img.get("uri") or img.get("resource_url") or ""


# ---------- HTML writer ----------

def write_index_html(items: List[Dict]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    COVERS_DIR.mkdir(parents=True, exist_ok=True)

    # Shared styling
    style = '''
<style>
  body { font-family: Arial, Helvetica, sans-serif; }
  .controls { position: sticky; top: 0; background: white; padding: 10px 0 12px 0; border-bottom: 1px solid #ddd; }
  .controls label { margin-right: 10px; }
  .controls input { width: 90px; padding: 4px 6px; }
  .controls .count { margin-left: 14px; color: #444; }
  .grid { display: flex; flex-wrap: wrap; gap: 18px; padding-top: 14px; }
  .card { display: inline-block; }
  .card img { display: block; width: auto; height: auto; }
  .meta { margin-top: 6px; max-width: 220px; }
  .meta a { text-decoration: none; color: inherit; }
  .meta a:hover { text-decoration: underline; }
</style>
'''.strip()

    # Compute year bounds (ignore 0/None)
    years = [int(it["year"]) for it in items if str(it.get("year") or "").isdigit() and int(it["year"]) > 0]
    min_year = min(years) if years else ""
    max_year = max(years) if years else ""

    # Index page (single page, all records)
    idx = OUTPUT_DIR / "index.html"
    lines: List[str] = [
        "<html><head>",
        "<meta charset='utf-8'>",
        style,
        "</head><body>",
        "<div class='controls'>",
        "<label>Min year <input id='minYear' type='number' inputmode='numeric'></label>",
        "<label>Max year <input id='maxYear' type='number' inputmode='numeric'></label>",
        "<span class='count' id='count'></span>",
        "</div>",
        "<div class='grid' id='grid'>",
    ]

    for it in items:
        img = it.get("img_local")
        artist = it.get("artist", "?")
        title = it.get("title", "?")
        year = it.get("year") or ""
        discogs_url = it.get("discogs_url")

        pic = f"<img src='{img}'>" if img else ""
        label = f"{year} — {artist} — {title}" if year else f"{artist} — {title}"
        if discogs_url:
            meta = f"<a href='{discogs_url}' target='_blank' rel='noopener noreferrer'>{label}</a>"
        else:
            meta = label

        year_attr = year if str(year).isdigit() else ""
        lines.append(f"<div class='card' data-year='{year_attr}'>")
        if pic:
            lines.append(pic)
        lines.append(f"<div class='meta'>{meta}</div>")
        lines.append("</div>")

    lines += [
        "</div>",
        "<script>",
        f"const minDefault = {json.dumps(min_year)};",
        f"const maxDefault = {json.dumps(max_year)};",
        "const minInput = document.getElementById('minYear');",
        "const maxInput = document.getElementById('maxYear');",
        "const countEl = document.getElementById('count');",
        "const cards = Array.from(document.querySelectorAll('.card'));",
        "minInput.value = minDefault;",
        "maxInput.value = maxDefault;",
        "function parseNum(v){ const n = parseInt(v, 10); return Number.isFinite(n) ? n : null; }",
        "function applyFilter(){",
        "  const minY = parseNum(minInput.value);",
        "  const maxY = parseNum(maxInput.value);",
        "  let shown = 0;",
        "  for (const c of cards){",
        "    const yRaw = c.getAttribute('data-year');",
        "    const y = parseNum(yRaw);",
        "    let ok = true;",
        "    if (minY !== null && y !== null && y < minY) ok = false;",
        "    if (maxY !== null && y !== null && y > maxY) ok = false;",
        "    // If a card has no year, hide it when any bound is set.",
        "    if ((minY !== null || maxY !== null) && y === null) ok = false;",
        "    c.style.display = ok ? 'inline-block' : 'none';",
        "    if (ok) shown += 1;",
        "  }",
        "  countEl.textContent = `${shown} / ${cards.length} shown`;",
        "}",
        "minInput.addEventListener('input', applyFilter);",
        "maxInput.addEventListener('input', applyFilter);",
        "applyFilter();",
        "</script>",
        "</body></html>",
    ]

    idx.write_text("\n".join(lines), encoding="utf-8")


# ---------- main ----------


def main() -> None:
    setup_logging()

    if not TOKEN or not USER:
        logging.error("DISCOGS_TOKEN or DISCOGS_USERNAME missing from environment.")
        logging.error("Make sure D:\\records\\.env exists and has DISCOGS_TOKEN / DISCOGS_USERNAME.")
        return

    client = Client()
    files = find_input_files()
    if not files:
        logging.warning("No input files found; nothing to do.")
        return

    cache: Dict[int, Dict] = {}
    all_items: List[Dict] = []

    COVERS_DIR.mkdir(parents=True, exist_ok=True)

    for path in files:
        # decade_label = infer_decade(path)  # no longer used (single-page index)

        if path.suffix.lower() == ".csv":
            rows = read_csv(path)
        else:
            rows = read_xlsx(path)

        for row in rows:
            rid = extract_release_id(row)
            if not rid:
                continue

            if rid not in cache:
                cache[rid] = client.get_release(rid)
                time.sleep(0.4)
            rel = cache.get(rid) or {}

            title = (rel.get("title") or row.get("title") or "").strip()
            arts = rel.get("artists") or []
            if arts:
                artist = ", ".join(a.get("name", "") for a in arts)
            else:
                artist = (row.get("artist") or "").strip()

            img_local = None
            url = choose_cover(rel)
            if url:
                fname = f"{rid}.jpg"
                dest = COVERS_DIR / fname
                if not dest.exists():
                    if download_image(url, dest):
                        logging.info("Saved cover %s for release %s", fname, rid)
                if dest.exists():
                    img_local = f"covers/{fname}"

            discogs_url = (rel.get("uri") or "").strip() or f"https://www.discogs.com/release/{rid}"
                        # Year for filtering (Discogs 'year' is typically an int)
            year_val = rel.get("year") or row.get("year") or row.get("release_year") or ""
            year_str = str(year_val).strip()
            year = int(year_str) if year_str.isdigit() else ""
            all_items.append({"year": year, "title": title, "artist": artist, "img_local": img_local, "discogs_url": discogs_url})
    # Sort: year asc (unknowns last), then artist/title
    all_items.sort(key=lambda x: (x.get('year') or 9999, (x.get('artist') or '').lower(), (x.get('title') or '').lower()))
    write_index_html(all_items)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        pass
